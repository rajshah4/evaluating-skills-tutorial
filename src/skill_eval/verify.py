from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .constants import get_task_config


DEPENDENCY_REQUIRED_KEYS = [
    "package",
    "version",
    "cve_id",
    "severity",
    "cvss_score",
    "fixed_version",
    "title",
    "url",
]


@dataclass
class VerificationResult:
    passed: bool
    message: str
    item_count: int


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _normalize_dependency_finding(finding: dict[str, Any]) -> dict[str, str]:
    normalized = {}
    for key in DEPENDENCY_REQUIRED_KEYS:
        value = finding.get(key)
        if value is None:
            raise ValueError(f"Missing required key: {key}")
        normalized[key] = str(value).strip()
    return normalized


def verify_dependency_audit(report_path: Path, expected_path: Path) -> VerificationResult:
    report = load_json(report_path)
    expected = load_json(expected_path)

    findings = report.get("findings")
    if not isinstance(findings, list):
        return VerificationResult(False, "report.json must contain a top-level 'findings' list", 0)

    try:
        actual = sorted(
            (_normalize_dependency_finding(item) for item in findings),
            key=lambda item: (item["package"], item["version"], item["cve_id"]),
        )
        target = sorted(
            (_normalize_dependency_finding(item) for item in expected["findings"]),
            key=lambda item: (item["package"], item["version"], item["cve_id"]),
        )
    except ValueError as exc:
        return VerificationResult(False, str(exc), len(findings))

    if actual != target:
        return VerificationResult(
            False,
            "report.json does not match the expected HIGH/CRITICAL vulnerability set",
            len(findings),
        )

    return VerificationResult(True, "report.json matches the expected benchmark findings", len(findings))


def verify_sec_financial_report(report_path: Path, expected_path: Path) -> VerificationResult:
    report = load_json(report_path)
    expected = load_json(expected_path)

    required_fields = expected["required_fields"]
    for field in required_fields:
        if field not in report:
            return VerificationResult(False, f"answers.json is missing required field: {field}", 0)

    for field, expected_value in expected["exact_values"].items():
        if report.get(field) != expected_value:
            return VerificationResult(
                False,
                f"answers.json field '{field}' did not match expected value",
                len(required_fields),
            )

    for field, spec in expected["numeric_values"].items():
        raw_value = report.get(field)
        try:
            actual_value = float(raw_value)
        except (TypeError, ValueError):
            return VerificationResult(
                False,
                f"answers.json field '{field}' must be numeric",
                len(required_fields),
            )

        expected_value = float(spec["value"])
        tolerance = float(spec["tolerance"])
        if abs(actual_value - expected_value) > tolerance:
            return VerificationResult(
                False,
                f"answers.json field '{field}' was outside tolerance",
                len(required_fields),
            )

    for field, substrings in expected.get("contains_values", {}).items():
        actual_value = report.get(field)
        if not isinstance(actual_value, str):
            return VerificationResult(
                False,
                f"answers.json field '{field}' must be a string",
                len(required_fields),
            )
        normalized = actual_value.lower()
        missing = [substring for substring in substrings if substring.lower() not in normalized]
        if missing:
            return VerificationResult(
                False,
                f"answers.json field '{field}' is missing expected content",
                len(required_fields),
            )

    return VerificationResult(
        True,
        "answers.json matches the expected financial extraction and calculations",
        len(required_fields),
    )


def verify_sales_pivot_analysis(report_path: Path, expected_path: Path) -> VerificationResult:
    if not report_path.exists():
        return VerificationResult(False, "result.xlsx was not created", 0)

    expected = load_json(expected_path)
    workbook = load_workbook(report_path, data_only=True)
    sheetnames = workbook.sheetnames

    for name in expected["required_sheets"]:
        if name not in sheetnames:
            return VerificationResult(False, f"result.xlsx is missing required sheet: {name}", 0)

    combined_sheet = workbook["CombinedData"]
    header_row = next(combined_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    required_columns = expected["combined_data"]["required_columns"]
    missing_columns = [column for column in required_columns if column not in headers]
    if missing_columns:
        return VerificationResult(False, "CombinedData is missing required columns", 0)

    header_index = {header: idx for idx, header in enumerate(headers)}
    actual_rows: dict[str, dict[str, Any]] = {}
    for row in combined_sheet.iter_rows(min_row=2, values_only=True):
        city = row[header_index["City"]]
        if city is None:
            continue
        city_name = str(city).strip()
        if not city_name:
            continue
        actual_rows[city_name] = {column: row[header_index[column]] for column in required_columns if column in header_index}

    expected_rows = expected["combined_data"]["rows"]
    for city, row_spec in expected_rows.items():
        if city not in actual_rows:
            return VerificationResult(False, f"CombinedData is missing row for {city}", len(expected_rows))
        actual_row = actual_rows[city]
        for field, expected_value in row_spec.items():
            actual_value = actual_row.get(field)
            if isinstance(expected_value, (int, float)):
                try:
                    numeric_actual = float(actual_value)
                except (TypeError, ValueError):
                    return VerificationResult(False, f"CombinedData.{city}.{field} must be numeric", len(expected_rows))
                if abs(numeric_actual - float(expected_value)) > 0.01:
                    return VerificationResult(False, f"CombinedData.{city}.{field} did not match expected value", len(expected_rows))
            else:
                if str(actual_value).strip() != str(expected_value):
                    return VerificationResult(False, f"CombinedData.{city}.{field} did not match expected value", len(expected_rows))

    summary_sheet = workbook["Summary"]
    label_value_pairs: dict[str, Any] = {}
    city_value_pairs: dict[str, float] = {}
    for row in summary_sheet.iter_rows(values_only=True):
        first = row[0] if len(row) > 0 else None
        second = row[1] if len(row) > 1 else None
        if first is None:
            continue
        label = str(first).strip()
        if not label:
            continue
        normalized = label.lower().replace("_", " ")
        if second is not None:
            if "total revenue" in normalized:
                label_value_pairs["Total Revenue"] = second
            elif "total population" in normalized:
                label_value_pairs["Total Population"] = second
            elif label in expected["summary"]["city_revenue_per_capita"]:
                try:
                    city_value_pairs[label] = float(second)
                except (TypeError, ValueError):
                    return VerificationResult(False, f"Summary value for {label} must be numeric", len(expected_rows))

    for label, expected_value in expected["summary"]["totals"].items():
        actual_value = label_value_pairs.get(label)
        try:
            numeric_actual = float(actual_value)
        except (TypeError, ValueError):
            return VerificationResult(False, f"Summary is missing {label}", len(expected_rows))
        if abs(numeric_actual - float(expected_value)) > 0.01:
            return VerificationResult(False, f"Summary {label} did not match expected value", len(expected_rows))

    for city, expected_value in expected["summary"]["city_revenue_per_capita"].items():
        actual_value = city_value_pairs.get(city)
        if actual_value is None:
            return VerificationResult(False, f"Summary is missing revenue per capita for {city}", len(expected_rows))
        if abs(actual_value - float(expected_value)) > 0.01:
            return VerificationResult(False, f"Summary revenue per capita for {city} did not match expected value", len(expected_rows))

    return VerificationResult(
        True,
        "result.xlsx contains the expected sheets and derived values",
        len(expected_rows) + len(expected["summary"]["totals"]) + len(expected["summary"]["city_revenue_per_capita"]),
    )


def verify_task_output(task: str, report_path: Path, expected_path: Path | None = None) -> VerificationResult:
    config = get_task_config(task)
    expected_path = expected_path or config.expected_output

    if task == "software-dependency-audit":
        return verify_dependency_audit(report_path, expected_path)
    if task == "sec-financial-report":
        return verify_sec_financial_report(report_path, expected_path)
    if task == "sales-pivot-analysis":
        return verify_sales_pivot_analysis(report_path, expected_path)
    raise ValueError(f"Unsupported task: {task}")
