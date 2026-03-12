from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from skill_eval.verification import VerificationResult, load_json


def verify(report_path: Path, expected_path: Path) -> VerificationResult:
    if not report_path.exists():
        return VerificationResult(False, "result.xlsx was not created", 0)

    expected = load_json(expected_path)
    workbook = load_workbook(report_path, data_only=True)
    sheetnames = workbook.sheetnames

    for name in expected["required_sheets"]:
        if name not in sheetnames:
            return VerificationResult(False, f"result.xlsx is missing required sheet: {name}", 0)

    combined_sheet = workbook["CombinedData"]
    header_row = next(combined_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    required_columns = expected["combined_data"]["required_columns"]
    missing_columns = [column for column in required_columns if column not in headers]
    if missing_columns:
        return VerificationResult(False, "CombinedData is missing required columns", 0)

    header_index = {header: idx for idx, header in enumerate(headers)}
    actual_rows: dict[str, dict[str, Any]] = {}
    for row in combined_sheet.iter_rows(min_row=2, values_only=True):
        city = row[header_index["City"]]
        if city is None:
            continue
        city_name = str(city).strip()
        if not city_name:
            continue
        actual_rows[city_name] = {column: row[header_index[column]] for column in required_columns if column in header_index}

    expected_rows = expected["combined_data"]["rows"]
    for city, row_spec in expected_rows.items():
        if city not in actual_rows:
            return VerificationResult(False, f"CombinedData is missing row for {city}", len(expected_rows))
        actual_row = actual_rows[city]
        for field, expected_value in row_spec.items():
            actual_value = actual_row.get(field)
            if isinstance(expected_value, (int, float)):
                try:
                    numeric_actual = float(actual_value)
                except (TypeError, ValueError):
                    return VerificationResult(False, f"CombinedData.{city}.{field} must be numeric", len(expected_rows))
                if abs(numeric_actual - float(expected_value)) > 0.01:
                    return VerificationResult(False, f"CombinedData.{city}.{field} did not match expected value", len(expected_rows))
            else:
                if str(actual_value).strip() != str(expected_value):
                    return VerificationResult(False, f"CombinedData.{city}.{field} did not match expected value", len(expected_rows))

    summary_sheet = workbook["Summary"]
    label_value_pairs: dict[str, Any] = {}
    city_value_pairs: dict[str, float] = {}
    for row in summary_sheet.iter_rows(values_only=True):
        first = row[0] if len(row) > 0 else None
        second = row[1] if len(row) > 1 else None
        if first is None:
            continue
        label = str(first).strip()
        if not label:
            continue
        normalized = label.lower().replace("_", " ")
        if second is not None:
            if "total revenue" in normalized:
                label_value_pairs["Total Revenue"] = second
            elif "total population" in normalized:
                label_value_pairs["Total Population"] = second
            elif label in expected["summary"]["city_revenue_per_capita"]:
                try:
                    city_value_pairs[label] = float(second)
                except (TypeError, ValueError):
                    return VerificationResult(False, f"Summary value for {label} must be numeric", len(expected_rows))

    for label, expected_value in expected["summary"]["totals"].items():
        actual_value = label_value_pairs.get(label)
        try:
            numeric_actual = float(actual_value)
        except (TypeError, ValueError):
            return VerificationResult(False, f"Summary is missing {label}", len(expected_rows))
        if abs(numeric_actual - float(expected_value)) > 0.01:
            return VerificationResult(False, f"Summary {label} did not match expected value", len(expected_rows))

    for city, expected_value in expected["summary"]["city_revenue_per_capita"].items():
        actual_value = city_value_pairs.get(city)
        if actual_value is None:
            return VerificationResult(False, f"Summary is missing revenue per capita for {city}", len(expected_rows))
        if abs(actual_value - float(expected_value)) > 0.01:
            return VerificationResult(False, f"Summary revenue per capita for {city} did not match expected value", len(expected_rows))

    return VerificationResult(
        True,
        "result.xlsx contains the expected sheets and derived values",
        len(expected_rows) + len(expected["summary"]["totals"]) + len(expected["summary"]["city_revenue_per_capita"]),
    )
